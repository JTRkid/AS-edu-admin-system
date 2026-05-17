import bcrypt
from datetime import datetime, timedelta
from django.utils import timezone
from django.contrib.auth.hashers import make_password
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from .models import User, Student, Teacher, StudentGroup, VerifyCode, OperationLog
from .serializers import (
    LoginSerializer, UserSerializer, StudentSerializer, StudentCreateSerializer,
    TeacherSerializer, TeacherCreateSerializer, ChangePasswordSerializer,
    ResetPasswordSerializer, PhoneBindSerializer, StudentGroupSerializer,
    OperationLogSerializer,
)


def get_tokens_for_user(user):
    refresh = RefreshToken.for_user(user)
    return {
        'access_token': str(refresh.access_token),
        'refresh_token': str(refresh),
    }


class LoginViewSet(viewsets.ViewSet):
    authentication_classes = []
    permission_classes = [permissions.AllowAny]

    @action(methods=['post'], detail=False)
    def login(self, request):
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']
        tokens = get_tokens_for_user(user)

        user_data = {
            'id': user.id,
            'name': user.name,
            'role': user.role,
        }
        if user.role == 'student':
            try:
                student = Student.objects.get(user=user)
                user_data['student_no'] = student.student_no
                user_data['class_name'] = student.class_name
            except Student.DoesNotExist:
                pass
        elif user.role in ('teacher', 'admin'):
            try:
                teacher = Teacher.objects.get(user=user)
                user_data['teacher_no'] = teacher.teacher_no
                user_data['department'] = teacher.department
                user_data['is_admin'] = teacher.is_admin
            except Teacher.DoesNotExist:
                pass

        return Response({
            'code': 200,
            'message': '登录成功',
            'data': {**tokens, 'user': user_data},
        })

    @action(methods=['post'], detail=False)
    def reset_password(self, request):
        serializer = ResetPasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']
        code = serializer.validated_data['code_obj']

        # Hash new password with bcrypt
        hashed = bcrypt.hashpw(
            serializer.validated_data['new_password'].encode('utf-8'),
            bcrypt.gensalt()
        ).decode('utf-8')
        user.password = hashed
        user.save()

        # Mark code as used
        code.used = True
        code.save()

        return Response({
            'code': 200,
            'message': '密码重置成功',
        })

    @action(methods=['post'], detail=False)
    def send_verify_code(self, request):
        """Send SMS verification code (simulated)"""
        phone = request.data.get('phone', '')
        if not phone:
            return Response({'code': 400, 'message': '请输入手机号'}, status=400)

        # In production, integrate with SMS service
        code = '123456'  # Simulated code

        VerifyCode.objects.create(
            phone=phone,
            code=code,
            expires_at=timezone.now() + timedelta(minutes=5),
        )

        return Response({
            'code': 200,
            'message': '验证码已发送',
            'data': {'code': code},  # Only for dev; remove in production
        })


class UserViewSet(viewsets.ViewSet):
    """当前用户信息管理"""

    @action(methods=['get'], detail=False)
    def me(self, request):
        user = request.user
        user_data = UserSerializer(user).data

        if user.role == 'student':
            try:
                student = Student.objects.get(user=user)
                user_data['student_no'] = student.student_no
                user_data['class_name'] = student.class_name
                user_data['group_id'] = student.group_id
            except Student.DoesNotExist:
                pass
        elif user.role in ('teacher', 'admin'):
            try:
                teacher = Teacher.objects.get(user=user)
                user_data['teacher_no'] = teacher.teacher_no
                user_data['department'] = teacher.department
                user_data['is_admin'] = teacher.is_admin
            except Teacher.DoesNotExist:
                pass

        return Response({'code': 200, 'data': user_data})

    @action(methods=['put'], detail=False)
    def change_password(self, request):
        serializer = ChangePasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        user = request.user
        old_pw = serializer.validated_data['old_password']
        new_pw = serializer.validated_data['new_password']

        if not bcrypt.checkpw(old_pw.encode('utf-8'), user.password.encode('utf-8')):
            if not user.check_password(old_pw):
                return Response({'code': 400, 'message': '旧密码错误'}, status=400)

        if old_pw == new_pw:
            return Response({'code': 400, 'message': '新密码不能与旧密码相同'}, status=400)

        user.password = bcrypt.hashpw(new_pw.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        user.save()

        return Response({'code': 200, 'message': '密码修改成功'})

    @action(methods=['post'], detail=False)
    def bind_phone(self, request):
        serializer = PhoneBindSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        phone = serializer.validated_data['phone']
        code = serializer.validated_data['verify_code']

        # Verify code
        vc = VerifyCode.objects.filter(
            phone=phone, code=code, used=False, expires_at__gt=timezone.now()
        ).first()
        if not vc:
            return Response({'code': 400, 'message': '验证码无效'}, status=400)
        if vc.attempt_count >= 3:
            return Response({'code': 400, 'message': '验证码尝试次数已达上限'}, status=400)

        user = request.user
        user.phone = phone
        user.phone_verified = True
        user.save()

        vc.used = True
        vc.save()

        return Response({'code': 200, 'message': '手机号绑定成功'})


class StudentViewSet(viewsets.ModelViewSet):
    """教师端学生管理"""
    queryset = Student.objects.select_related('user').all()
    serializer_class = StudentSerializer
    filterset_fields = ['class_name']
    search_fields = ['student_no', 'user__name', 'class_name']

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated()]

    def get_serializer_class(self):
        if self.action == 'create':
            return StudentCreateSerializer
        return StudentSerializer

    def create(self, request, *args, **kwargs):
        serializer = StudentCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        student = serializer.save()
        return Response({
            'code': 200,
            'message': '学生添加成功',
            'data': StudentSerializer(student).data,
        }, status=201)

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response({'code': 200, 'data': serializer.data})

    @action(methods=['post'], detail=True)
    def reset_password(self, request, pk=None):
        student = self.get_object()
        new_pw = request.data.get('password', '123456')
        student.user.password = bcrypt.hashpw(
            new_pw.encode('utf-8'), bcrypt.gensalt()
        ).decode('utf-8')
        student.user.save()
        return Response({'code': 200, 'message': '密码已重置'})

    @action(methods=['post'], detail=False)
    def batch_import(self, request):
        """批量导入学生（Excel）"""
        import re
        import openpyxl
        from io import BytesIO

        file = request.FILES.get('file')
        if not file:
            return Response({'code': 400, 'message': '请上传文件'}, status=400)

        wb = openpyxl.load_workbook(BytesIO(file.read()))
        ws = wb.active

        created = 0
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue
            student_no = str(row[0]).strip() if row[0] else ''
            name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            class_name = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            password = str(row[3]).strip() if len(row) > 3 and row[3] else '123456'

            # 验证学号：13位数字
            if not re.match(r'^\d{13}$', student_no):
                errors.append(f'第{row_idx}行 {student_no}: 学号必须为13位数字')
                continue

            # 验证班级号：9位数字
            if not re.match(r'^\d{9}$', class_name):
                errors.append(f'第{row_idx}行 {student_no}: 班级号必须为9位数字')
                continue

            # 检查学号是否已被使用（包括其他角色）
            if User.objects.filter(username=student_no).exists():
                errors.append(f'第{row_idx}行 {student_no}: 该学号已被使用')
                continue

            if Student.objects.filter(student_no=student_no).exists():
                errors.append(f'第{row_idx}行 {student_no}: 已存在')
                continue

            try:
                hashed = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                user = User.objects.create(username=student_no, name=name, password=hashed, role='student')
                Student.objects.create(user=user, student_no=student_no, class_name=class_name)
                created += 1
            except Exception as e:
                errors.append(f'第{row_idx}行 {student_no}: {str(e)}')

        return Response({
            'code': 200,
            'message': f'导入完成，成功{created}条',
            'data': {'created': created, 'errors': errors},
        })

    @action(methods=['get'], detail=False)
    def export_excel(self, request):
        """导出学生名单为Excel"""
        import io
        import openpyxl
        from django.http import HttpResponse

        queryset = self.filter_queryset(self.get_queryset())
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '学生名单'
        ws.append(['学号', '姓名', '班级', '状态'])
        for s in queryset:
            ws.append([s.student_no, s.user.name, s.class_name, '正常' if s.user.is_active else '禁用'])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=students.xlsx'
        return response

    @action(methods=['get'], detail=False)
    def download_template(self, request):
        """下载学生导入模板"""
        import io
        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '学生导入模板'
        ws.append(['学号', '姓名', '班级', '密码'])
        ws.append(['2023001', '张三', '计算机1班', '123456'])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=student_template.xlsx'
        return response


class TeacherViewSet(viewsets.ModelViewSet):
    """管理员端教师管理"""
    queryset = Teacher.objects.select_related('user').all()
    serializer_class = TeacherSerializer
    search_fields = ['teacher_no', 'user__name', 'department']

    def get_serializer_class(self):
        if self.action == 'create':
            return TeacherCreateSerializer
        return TeacherSerializer

    def create(self, request, *args, **kwargs):
        serializer = TeacherCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        teacher = serializer.save()
        return Response({
            'code': 200,
            'message': '教师添加成功',
            'data': TeacherSerializer(teacher).data,
        }, status=201)

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response({'code': 200, 'data': serializer.data})

    @action(methods=['post'], detail=True)
    def set_admin(self, request, pk=None):
        teacher = self.get_object()
        teacher.is_admin = True
        teacher.user.role = 'admin'
        teacher.user.save()
        teacher.save()

        OperationLog.objects.create(
            user=request.user, action='设置管理员',
            target=f'{teacher.teacher_no}', detail='提升为管理员',
        )
        return Response({'code': 200, 'message': '已设为管理员'})

    @action(methods=['post'], detail=True)
    def revoke_admin(self, request, pk=None):
        teacher = self.get_object()
        teacher.is_admin = False
        teacher.user.role = 'teacher'
        teacher.user.save()
        teacher.save()

        OperationLog.objects.create(
            user=request.user, action='撤销管理员',
            target=f'{teacher.teacher_no}', detail='降级为普通教师',
        )
        return Response({'code': 200, 'message': '已撤销管理员'})

    @action(methods=['post'], detail=True)
    def toggle_active(self, request, pk=None):
        teacher = self.get_object()
        teacher.user.is_active = not teacher.user.is_active
        teacher.user.save()
        status_text = '启用' if teacher.user.is_active else '禁用'
        return Response({'code': 200, 'message': f'教师账号已{status_text}'})

    @action(methods=['post'], detail=True)
    def reset_password(self, request, pk=None):
        teacher = self.get_object()
        new_pw = request.data.get('password', '123456')
        teacher.user.password = bcrypt.hashpw(
            new_pw.encode('utf-8'), bcrypt.gensalt()
        ).decode('utf-8')
        teacher.user.save()
        return Response({'code': 200, 'message': '密码已重置'})


class StudentGroupViewSet(viewsets.ModelViewSet):
    queryset = StudentGroup.objects.all()
    serializer_class = StudentGroupSerializer

    def perform_create(self, serializer):
        serializer.save(teacher=self.request.user)


class OperationLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = OperationLog.objects.select_related('user').all().order_by('-created_at')
    serializer_class = OperationLogSerializer
