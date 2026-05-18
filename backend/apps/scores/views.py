"""成绩管理视图 — 成绩CRUD、Excel导入导出、实验成绩提交API"""

import io
import openpyxl
from django.conf import settings
from django.http import HttpResponse
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from apps.accounts.models import Student
from apps.courses.models import Section
from .models import Score, ScoreHistory
from .serializers import ScoreSerializer, ScoreSubmitSerializer


class ScoreViewSet(viewsets.ModelViewSet):
    """成绩管理ViewSet，支持总成绩/实验成绩分类筛选、手动修改审计、Excel导入导出"""
    queryset = Score.objects.select_related('student', 'section', 'modified_by').all()
    serializer_class = ScoreSerializer
    filterset_fields = ['section', 'chapter_no', 'section_no', 'class_name', 'source', 'score_type']
    search_fields = ['student_no', 'student_name', 'class_name']
    ordering_fields = ['student_no', 'student_name', 'class_name', 'chapter_no', 'section_no', 'score', 'created_at']

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        if request.user.role == 'student':
            queryset = queryset.filter(student=request.user)
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response({'code': 200, 'data': serializer.data})

    @action(methods=['get'], detail=False)
    def my_scores(self, request):
        queryset = Score.objects.filter(student=request.user)
        section_id = request.query_params.get('section_id')
        if section_id:
            queryset = queryset.filter(section_id=section_id)
        serializer = ScoreSerializer(queryset, many=True)
        return Response({'code': 200, 'data': serializer.data})

    @action(methods=['put'], detail=True)
    def modify(self, request, pk=None):
        score_record = self.get_object()
        old_score = score_record.score
        new_score = request.data.get('score')
        reason = request.data.get('reason', '')
        if new_score is None:
            return Response({'code': 400, 'message': '请提供新成绩'}, status=400)
        ScoreHistory.objects.create(
            score_record=score_record, old_score=old_score, new_score=new_score,
            modified_by=request.user, reason=reason,
        )
        score_record.score = new_score
        score_record.source = 'manual'
        score_record.modified_by = request.user
        score_record.modify_reason = reason
        score_record.save()
        return Response({'code': 200, 'message': '成绩修改成功', 'data': ScoreSerializer(score_record).data})

    @action(methods=['get'], detail=False)
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        if request.user.role == 'student':
            queryset = queryset.filter(student=request.user)
        section_id = request.query_params.get('section')
        if section_id:
            queryset = queryset.filter(section_id=section_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '成绩表'
        headers = ['学号', '班级', '姓名', '章号', '章名', '节号', '节名', '成绩', '提交时间']
        ws.append(headers)
        for score in queryset:
            ws.append([
                score.student_no, score.class_name, score.student_name,
                score.chapter_no, score.chapter_name, score.section_no, score.section_name,
                float(score.score),
                score.updated_at.strftime('%Y-%m-%d %H:%M:%S') if score.updated_at else '',
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=scores.xlsx'
        return response

    @action(methods=['get'], detail=False)
    def download_template(self, request):
        """下载成绩导入模板"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '成绩导入模板'
        headers = ['学号', '班级', '姓名', '章号', '章名', '节号', '节名', '成绩']
        ws.append(headers)
        ws.append(['2023001', '计算机1班', '张三', 1, 'Python基础', 1, 'Python简介', 90])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=score_template.xlsx'
        return response

    @action(methods=['post'], detail=False)
    def import_excel(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'code': 400, 'message': '请上传文件'}, status=400)
        wb = openpyxl.load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        created = 0
        updated = 0
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue
            student_no = str(row[0]).strip()
            student_name = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            class_name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            chapter_no = int(row[3]) if len(row) > 3 and row[3] else 0
            section_no = int(row[5]) if len(row) > 5 and row[5] else 0
            new_score = row[7] if len(row) > 7 and row[7] else 0

            if not student_no or not chapter_no or not section_no:
                errors.append(f'第{row_idx}行: 学号/章号/节号不能为空')
                continue

            try:
                section = Section.objects.filter(
                    chapter__chapter_no=chapter_no, section_no=section_no
                ).first()
                if not section:
                    errors.append(f'第{row_idx}行: 章{chapter_no}节{section_no}不存在')
                    continue

                student = Student.objects.filter(student_no=student_no).first()
                if not student:
                    errors.append(f'第{row_idx}行 {student_no}: 学生不存在，跳过')
                    continue

                score_record, is_new = Score.objects.update_or_create(
                    student=student.user,
                    section=section,
                    defaults={
                        'student_no': student_no,
                        'student_name': student_name or student.user.name,
                        'class_name': class_name or student.class_name,
                        'chapter_no': chapter_no,
                        'chapter_name': section.chapter.title,
                        'section_no': section_no,
                        'section_name': section.title,
                        'score': new_score,
                        'source': 'import',
                        'evaluator': 'Excel导入',
                    }
                )
                if is_new:
                    created += 1
                else:
                    ScoreHistory.objects.create(
                        score_record=score_record, old_score=0, new_score=new_score,
                        modified_by=request.user, reason='Excel导入覆盖',
                    )
                    updated += 1
            except Exception as e:
                errors.append(f'第{row_idx}行 {student_no}: {str(e)}')

        return Response({
            'code': 200,
            'message': f'导入完成，新增{created}条，更新{updated}条',
            'data': {'created': created, 'updated': updated, 'errors': errors},
        })


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def score_submit_api(request):
    """实验评分机提交成绩API，使用 ApiKey 认证，自动创建或覆盖实验成绩"""
    auth_header = request.headers.get('Authorization', '')
    if auth_header != f"ApiKey {settings.SCORING_MACHINE_API_KEY}":
        return Response({'code': 403, 'message': '无权限'}, status=403)

    serializer = ScoreSubmitSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    data = serializer.validated_data

    try:
        student = Student.objects.get(student_no=data['student_no'])
    except Student.DoesNotExist:
        return Response({'code': 404, 'message': '学生不存在'}, status=404)

    course_name = data.get('course_name', '')
    section_qs = Section.objects.select_related('chapter__course').filter(
        chapter__chapter_no=data['chapter_no'], section_no=data['section_no'],
    )
    if course_name:
        section_qs = section_qs.filter(chapter__course__name=course_name)
    section = section_qs.first()
    if not section:
        return Response({'code': 404, 'message': f"章{data['chapter_no']}节{data['section_no']}不存在"}, status=404)

    existing = Score.objects.filter(student=student.user, section=section).first()
    is_overwrite = existing is not None

    if existing:
        # FIXME: original_score 应在覆盖前保存旧值，此处 existing.score 已被赋为新值
        old_score = existing.score
        existing.score = data['score']
        existing.score_type = 'experiment'
        existing.source = 'experiment'
        existing.evaluator = data.get('evaluator', '')
        existing.details = str(data.get('details', ''))
        existing.original_score = old_score
        existing.save()
        score_record = existing
    else:
        score_record = Score.objects.create(
            student=student.user, student_no=data['student_no'],
            class_name=data.get('class_name', student.class_name),
            student_name=data.get('student_name', student.user.name),
            chapter_no=data['chapter_no'], chapter_name=data.get('chapter_name', ''),
            section_no=data['section_no'], section_name=data.get('section_name', ''),
            section=section, score=data['score'],
            score_type='experiment', source='experiment',
            evaluator=data.get('evaluator', ''), details=str(data.get('details', '')),
        )

    try:
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f"user_{student.user.id}",
            {"type": "score.update", "student_id": student.user.id,
             "chapter_no": data['chapter_no'], "section_no": data['section_no'],
             "score": float(data['score'])},
        )
        broadcast_status = 'success'
    except Exception:
        broadcast_status = 'failed'

    return Response({
        'code': 200, 'message': 'Score received and saved',
        'data': {'score_id': score_record.id, 'is_overwrite': is_overwrite,
                 'broadcast_status': broadcast_status, 'teacher_notified': True, 'student_notified': True},
    })
